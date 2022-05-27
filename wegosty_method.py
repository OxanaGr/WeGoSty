from selenium import webdriver
import wegosty_locators as locators
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from time import sleep
import datetime
from selenium.webdriver.common.by import By
import sys
import random
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.styles import fonts

s = Service(executable_path='chromedriver.exe')
driver = webdriver.Chrome(service=s)

def setUp():
    print(f'lanch {locators.app}')
    print('___________________________________________')
    driver.maximize_window()
    driver.implicitly_wait(30)
    driver.get(locators.wegosty_url)
    if driver.current_url == locators.wegosty_homepage_url:
        print(f' WeGoStudy URL: {driver.current_url}')
        print(f'__________________Test started successfully at {datetime.datetime.now()}__________________')
    # assert driver.current_url == locators.wegosty_homepage_url


def login():
    sleep(0.3)
    # driver.find_element(By.CSS_SELECTOR, '.toast-message').click()
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'LOGIN').click()
    sleep(0.3)
    # driver.find_element(By.ID, 'user_email').send_keys('chris.velasco78@gmail.com')
    driver.find_element(By.ID, 'user_email').send_keys(locators.partner_name)
    sleep(0.3)
    # driver.find_element(By.ID, 'user_password').send_keys(f'123cctb')
    driver.find_element(By.ID, 'user_password').send_keys(locators.partner_pass)
    sleep(0.3)
    # driver.find_element(By.ID, 'user_password').send_keys(f'123cctb{Keys.ENTER}')
    driver.find_element(By.CSS_SELECTOR, 'input[value="SIGN IN"]').click()
    driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]').click()
    sleep(2)
    # driver.find_element(By.XPATH, "//input[@value='SIGN IN']").click()
    assert driver.find_element(By.CSS_SELECTOR, 'img[alt="Partner"]').is_displayed()
    print(f'________Signed in successfully! at {datetime.datetime.now()}_________')


def log_out():
    sleep(0.8)
    driver.find_element(By.XPATH, '//div[@id="toast-container"]').click()
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'span[class="my-auto mr-2 pf-name"]').click()
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'Log out').click()
    if driver.find_element(By.XPATH, '//div[@id="toast-container"]').is_displayed():
        print(f'________Singed out successfully! at {datetime.datetime.now()}_________')
    # driver.find_element(By.XPATH, '//div[@id="toast-container"]').click()


def create_new_student():
    sleep(1)
    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(1)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(1)
    driver.find_element(By.LINK_TEXT, 'Create New Student').click()
    sleep(1)
    # __________________Personal_Information______________________________
    driver.find_element(By.ID, 'user_student_detail_attributes_first_name').send_keys(locators.first_name)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_last_name').send_keys(locators.last_name)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_preferred_name').send_keys(locators.first_name)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').send_keys(Keys.ENTER)
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').clear()
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'input[placeholder="Date of Birth on passport"]').send_keys('20000202')
    # driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').send_keys(locators.date_of_birth)
    # driver.find_element(By.ID, 'user_student_detail_attributes_birth_date').send_keys(f'{locators.date_of_birth}{Keys.ENTER}')
    sleep(0.3)
    driver.find_element(By.ID, 'phone_number').send_keys(locators.phone_number)
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'span[title="Country of citizenship"]').click()
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'input[role="searchbox"]').send_keys(f'Cameroon{Keys.ENTER}')
    sleep(0.3)
    driver.find_element(By.ID,'user_student_detail_attributes_passport_number').send_keys(locators.passport_number)
    sleep(0.3)
    # __________________Contact Information______________________________
    driver.find_element(By.ID, 'user_student_detail_attributes_address_attributes_mailing_address').send_keys(locators.mailing_address)
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'Country').click()
    sleep(0.3)
    driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_address_attributes_country_chosen"]/div/div/input').send_keys(f'Canada{Keys.ENTER}')
    # driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_address_attributes_country_chosen"]/div/ul/li[40]').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Province/State').click()
    sleep(0.6)
    driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_address_attributes_state_chosen"]/div/ul/li[3]').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'City').click()
    sleep(0.6)
    driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_address_attributes_city_chosen"]/div/ul/li[30]').click()
    sleep(0.6)
    driver.find_element(By.ID, 'user_student_detail_attributes_address_attributes_zip_code').send_keys(locators.postal_code)
    sleep(0.6)
    driver.find_element(By.ID, 'user_email').send_keys(locators.user_email)
    sleep(0.6)
    # __________________Education Information______________________________
    driver.find_element(By.LINK_TEXT, 'Credentials').click()
    sleep(0.6)
    driver.find_element(By.XPATH, '//*[@id="user_student_detail_attributes_user_educations_attributes_0_credentials_chosen"]/div/ul/li[3]').click()
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_user_educations_attributes_0_school_name').send_keys('CCTB')
    sleep(0.3)
    driver.find_element(By.ID, 'user_student_detail_attributes_user_educations_attributes_0_program').send_keys('SQTA')
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'GPA Scale').click()
    sleep(0.3)
    driver.find_element(By.XPATH, '//div[@id="user_student_detail_attributes_user_educations_attributes_0_gpa_scale_chosen"]/div/ul/li[5]').click()
    sleep(0.3)
    # __________________________________________
    driver.find_element(By.ID, 'user_student_detail_attributes_user_educations_attributes_0_gpa').send_keys(f'4.0')
    sleep(0.3)
    driver.find_element(By.XPATH, '//input[@value="Save"]').submit()
    # __________________________________________

    print(f'The new student {locators.first_name} {locators.last_name} is created.')
    print(f'Personal Information: Date Of Birth - {locators.date_of_birth}, Passport Number - {locators.passport_number}, Citizenship - Canada, Phone Number - {locators.phone_number}')
    print(f'Contact Information: Mailing Address - {locators.mailing_address}, Country - Canada, State - BC, City -  Burnaby, Postal Code - {locators.postal_code}')
    # print(f'Credential -, School Name -, Program -, GPA Scale -, Scored GPA - ')
    print('___________________________________________')


def create_application():
    l = 0.6
    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(l)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(l)
    driver.find_element(By.LINK_TEXT, 'Create Application').click()
    sleep(l)
    # ----------------Course Information----------------#
    driver.find_element(By.LINK_TEXT, 'Select School').click()
    sleep(l)
    driver.find_element(By.CSS_SELECTOR, 'input[class="chosen-search-input"]').send_keys('British Columbia')
    sleep(l)
    driver.find_element(By.XPATH, '//*[@id="admission_institute_detail_id_chosen"]/a/span').click()
    # driver.find_element(By.LINK_TEXT, 'British Columbia Institute of Technology').click()
    sleep(l)
    driver.find_element(By.LINK_TEXT, "Select Course").click()
    driver.find_element(By.XPATH, '//div[@id="admdission_institute_program_id_chosen"]/div/ul/li[2]').click()
    sleep(l)
    driver.find_element(By.LINK_TEXT, "Select Starting Semester").click()
    driver.find_element(By.XPATH, '//div[@id="admission_starting_semester_chosen"]/div/ul/li[2]').click()
    sleep(l)
    driver.find_element(By.LINK_TEXT, "Select Start Day").click()
    driver.find_element(By.XPATH, '//div[@id="admission_start_day_chosen"]/div/ul/li[2]').click()
    sleep(l)
    driver.find_element(By.LINK_TEXT, "Select Year").click()
    driver.find_element(By.XPATH, '//div[@id="admission_start_year_chosen"]/div/ul/li[2]').click()
    sleep(l)
    # ----------------Personal Information----------------#
    driver.find_element(By.XPATH, '//input[@id="admission_date_of_birth"]').clear()
    driver.find_element(By.XPATH, '//input[@id="admission_date_of_birth"]').click()
    i = 1
    while i <= 50:
        driver.find_element(By.XPATH, '//th[@class="next"]').click()
        i += 1
    sleep(l)
    driver.find_element(By.XPATH, '//div[@class="datepicker-days"]').click()
    sleep(l)
    driver.find_element(By.ID, "admission_passport_number").clear()
    driver.find_element(By.ID, "admission_passport_number").send_keys(locators.passport_number)
    driver.find_element(By.ID, "admission_cambrian_id").send_keys(locators.college_ID)
    driver.find_element(By.ID, "admission_study_permit_number").send_keys(locators.student_permit_number)
    driver.find_element(By.ID, "admission_first_language").send_keys(locators.first_language)
    driver.find_element(By.XPATH, '//input[@id="admission_date_of_admission"]').clear()
    sleep(l)
    # driver.find_element(By.XPATH, '//input[@id="admission_date_of_admission"]').send_keys('05-01-2010')
    driver.find_element(By.XPATH, '//input[@id="admission_date_of_admission"]').click()
    sleep(l)
    i = 1
    while i <= 3:
        driver.find_element(By.XPATH, '//th[@class="next"]').click()
        # driver.find_element(By.XPATH,'//body[1]/div[4]/div[1]/table[1]/thead[1]/tr[1]/th[1]').click()
        i += 1
    sleep(l)
    driver.find_element(By.XPATH, '//div[@class="datepicker-days"]').click()
    sleep(l)
    driver.find_element(By.XPATH, "//input[@id='admission_gender_male']").click()
    driver.find_element(By.XPATH, "//input[@id='admission_stay_type_private_residence']").click()
    driver.find_element(By.XPATH, "//input[@value='Save']").click()

def My_Referrals():
    # ___________________________________________________________
    from openpyxl import Workbook
    import openpyxl as O
    Exel_file = "C:\\Users\\ks\\Desktop\\OXANA\\QA\\wgsty\\WEG_test_result.xlsx"
    wb = O.load_workbook(Exel_file)
    ws = wb.active
    # ______________________________________________________________
    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'Referrals').click()
    sleep(0.3)
    driver.find_element(By.LINK_TEXT, 'Create Referral').click()
    sleep(0.3)
    # ______________________________________________________________________
    # driver.find_element(By.NAME, 'referral[first_name]').send_keys(ws['A11'].value)
    # sleep(0.3)
    # driver.find_element(By.NAME, 'referral[last_name]').send_keys(ws['B11'].value)
    # sleep(0.3)
    # driver.find_element(By.NAME, 'referral[email_id]').send_keys(ws['G11'].value)
    # _____________________________________________________________________
    driver.find_element(By.NAME, 'referral[first_name]').send_keys(locators.first_name)
    sleep(0.3)
    driver.find_element(By.NAME, 'referral[last_name]').send_keys(locators.last_name)
    sleep(0.3)
    driver.find_element(By.NAME, 'referral[email_id]').send_keys(locators.user_email)
    # _______________________________________________________________________
    sleep(0.3)
    driver.find_element(By.NAME, 'school_id').send_keys(f'Douglas College{Keys.ENTER}')
    sleep(0.3)
    driver.find_element(By.NAME, 'program_id').send_keys(f'Basic Musicianship Certificate{Keys.ENTER}')
    sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, 'input[value="Save"]').click()
    sleep(0.3)
    assert driver.find_element(By.XPATH, '//*[@id="referral-thank-you-modal"]'). is_displayed()
    sleep(0.3)
    notice = driver.find_element(By.XPATH, '//*[@id="wrapper"]')
    sleep(0.3)
    print(f'The confirmation of creation  referral  - is displayed')
    print('___________________________________________________________________________')
    # driver.find_element(By.CSS_SELECTOR, 'button[class="btn btn-default"]').click()


def search_student():
    # from openpyxl import Workbook
    import openpyxl as O
    Exel_file = "C:\\Users\\ks\\Desktop\\OXANA\\QA\\wgsty\\WEG_test_result.xlsx"
    wb = O.load_workbook(Exel_file)
    ws = wb.active

    driver.find_element(By.LINK_TEXT, 'My WeGoStudy').click()
    sleep(0.6)
    driver.find_element(By.LINK_TEXT, 'Students').click()
    sleep(0.6)
    # driver.find_element(By.ID, 'search').send_keys(locators.first_name, locators.last_name)
    driver.find_element(By.ID, 'search').send_keys (ws['A5'].value,' ', ws['B5'].value)
    sleep(0.6)
    driver.find_element(By.NAME, 'commit').click()


def tearDown():
    if driver is not None:
        print(f'_________________Test finished successfully at {datetime.datetime.now()}_________________')
        sleep(2)
        driver.close()
        driver.quit()

        logger('deleted')

def logger(action: object):
    old_instance = sys.stdout
    log_file = open('message.log', 'a')
    sys.stdout = log_file
    print(f'{locators.user_email}\t'
          f'{locators.first_name}\t'
          f'{locators.last_name}\t'
          f'{locators.mailing_address}\t'
          f'{locators.postal_code}\t'
          f'{locators.phone_number}\t'
          f'{datetime.datetime.now()}\t'
          f'{action}')
    sys.stdout = old_instance
    log_file.close()


def xlsx_data():
    from  openpyxl import Workbook
    import openpyxl as O
    Exel_file = "xxxxxxxxxxxxxxxxxxx"
    wb = O.load_workbook(Exel_file)
    ws = wb.active
    ws.append([f'{locators.first_name}', f'{locators.last_name}', f'{locators.passport_number}', f'{locators.mailing_address}',
               f'{locators.postal_code}',f'{locators.phone_number}', f'{locators.user_email}', f'{datetime.datetime.now()}',
               "test result"])
    wb.save(Exel_file)
# _______________________________________________________
#     ws['A5'] = f'{locators.first_name}'
#     ws['B5'] = f'{locators.last_name}'
#     ws['C5'] = f'{locators.passport_number}'
#     ws['D5'] = f'{locators.mailing_address}'
#     ws['E5'] = f'{locators.postal_code}'
#     ws['F5'] = f'{locators.phone_number}'
#     ws['G5'] = f'{locators.user_email}'
#     ws['H5'] = f'{datetime.datetime.now()}'
#     ws['I5'] = "test result"
#     wb.save(Exel_file)
# _________________________________________________



setUp()
login()
create_new_student()
# create_application()
# search_student()
# delete_studnt()
xlsx_data()
# My_Referrals()
# log_out()
# tearDown()
